import sys
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMessageBox, QAction, QFileDialog
from screens import *
import os
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from connectionDB import DbManager
from control_page import *
from helper_function import *

class App(MainWindow):
    def __init__(self):
        super().__init__()
        self.connection = DbManager()
        get_out = QAction("Quit", self)
        get_out.triggered.connect(self.closeEvent)



        # Triggered signals
        self.ui.act_branch.triggered.connect(lambda: self.open(BranchWindow))
        self.ui.act_cause.triggered.connect(lambda: self.open(CauseWindow))
        self.ui.act_class.triggered.connect(lambda: self.open(ClassWindow))
        self.ui.act_add_class_to_parent.triggered.connect(lambda: self.open(AddClassWindow))
        self.ui.act_degree.triggered.connect(lambda: self.open(DegreeWindow))
        self.ui.act_department.triggered.connect(lambda: self.open(DepartmentWindow))
        self.ui.act_disconnection.triggered.connect(lambda: self.open(DisconnectionWindow))
        self.ui.act_institutation_info.triggered.connect(lambda: self.open(InstitutionWindow))
        self.ui.act_neighborhood.triggered.connect(lambda: self.open(NeighborhoodWindow))
        self.ui.act_parent_class.triggered.connect(lambda: self.open(ParentClassWindow))
        self.ui.act_staff_branch.triggered.connect(lambda: self.open(StaffBranchWindow))
        self.ui.act_staff.triggered.connect(lambda: self.open(StaffWindow))
        self.ui.act_pull_sutudent.triggered.connect(lambda: self.open(StudentFromArchiveWindow))
        self.ui.act_student.triggered.connect(lambda: self.open(StudentWindow))
        self.ui.act_workplace.triggered.connect(lambda: self.open(WorkplaceWindow))
        self.ui.act_staff_workplace.triggered.connect(lambda: self.open(AddWorkplaceToStaff))

        self.ui.act_load_workplaces.triggered.connect(self.load_workplaces_file)
        self.ui.act_load_students.triggered.connect(self.load_students_file)
        self.ui.act_download_sample_workplaces.triggered.connect(self.save_excel_for_wp)
        self.ui.act_download_sample_students.triggered.connect(self.save_excel_file_for_student)

    def open(self, param_class):
        self.inst = param_class()
        self.inst.setWindowModality(Qt.ApplicationModal)
        self.inst.show()

    def read_wp_excel_file(self, path):
        loaded_wp_list = [i[0] for i in self.connection.selector(f"""SELECT name FROM workplaces""")]
        wb = load_workbook(path)
        active_page = wb['isletme']

        new_wp_list = []
        for row in range(2, 10000):
            if active_page.cell(row=row, column=1).value is None:
                break
            else:
                wp_name = tr_capitalize(stripper(str(active_page.cell(row, 1).value)))

                wp_boss = ""
                if active_page.cell(row, 2).value is not None:
                    wp_boss = tr_capitalize(stripper(str(active_page.cell(row, 2).value)))

                depart_name = active_page.cell(row, 3).value

                wp_mi = ""
                if active_page.cell(row, 4).value is not None:
                    wp_mi = tr_capitalize(stripper(str(active_page.cell(row, 4).value)))

                neigh_name = active_page.cell(row, 5).value

                add1 = ""
                if active_page.cell(row, 6).value is not None:
                    add1 = str(active_page.cell(row, 6).value)

                add2 = ""
                if active_page.cell(row, 7).value is not None:
                    add2 = str(active_page.cell(row, 7).value)

                tel1 = ""
                if active_page.cell(row, 8).value is not None:
                    tel1 = str(active_page.cell(row, 8).value)

                tel2 = ""
                if active_page.cell(row, 9).value is not None:
                    tel2 = str(active_page.cell(row, 9).value)

                email = ""
                if active_page.cell(row, 10).value is not None:
                    email = str(active_page.cell(row, 10).value)

                gov_ins = active_page.cell(row, 11).value

                control_list = [general_name_control(wp_name, loaded_wp_list), name_control(wp_boss), name_control(wp_mi), depart_name is not None, neigh_name is not None, gov_ins is not None]
                if False not in control_list:
                    depart_id = self.connection.find(f"""SELECT id FROM departments WHERE name="{depart_name}" """)
                    neigh_id = self.connection.find(f"""SELECT id FROM neighborhoods WHERE name="{neigh_name}" """)
                    new_wp_list.append((wp_name, wp_boss, depart_id, wp_mi, neigh_id, add1, add2, tel1, tel2, email, gov_ins))
                    loaded_wp_list.append(wp_name)
                else:
                    QMessageBox.warning(None, "Uyarı",  str(row-1) + ". satırda hata var!\ndüzeltip yeniden yükleyiniz.")
                    new_wp_list.clear()
                    break
        if len(new_wp_list) > 0:
            sql = "INSERT INTO workplaces(name, boss, departmentID, master_instructive, neighborhoodID, street, address_number, phone_number1, phone_number2, email, government_contribution)" \
                 "VALUES(?,?,?,?,?,?,?,?,?,?,?)"
            self.connection.poly_recorder(sql, new_wp_list)

    def read_student_excel_file(self, path):
        active_student_id_list = [i[0] for i in self.connection.selector(f"""SELECT id_number FROM students WHERE is_active = "E" """)]
        passive_student_id_list = [i[0] for i in self.connection.selector(f"""SELECT id_number FROM students WHERE is_active = "H" """)]
        wb = load_workbook(path)
        active_page = wb['Ogrenciler']

        new_student_list = []
        for row in range(2, 10000):
            if active_page.cell(row=row, column=1).value is None:
                break
            else:
                id_number = str(active_page.cell(row, 1).value)

                name = ""
                if active_page.cell(row, 2).value is not None:
                    name = tr_capitalize(stripper(str(active_page.cell(row, 2).value)))

                surname = ""
                if active_page.cell(row, 3).value is not None:
                    surname = tr_capitalize(stripper(str(active_page.cell(row, 3).value)))

                number = ""
                if active_page.cell(row, 4).value is not None:
                    number = tr_capitalize(stripper(str(active_page.cell(row, 4).value)))

                branch_name = active_page.cell(row, 5).value

                class_name = active_page.cell(row, 6).value
                if class_name is None:
                    class_name = ""

                recorded_date = ""
                if active_page.cell(row, 7).value is not None:
                    recorded_date = str(active_page.cell(row, 7).value)[:10]

                father_name = ""
                if active_page.cell(row, 8).value is not None:
                    father_name = tr_capitalize(stripper(str(active_page.cell(row, 8).value)))

                mother_name = ""
                if active_page.cell(row, 9).value is not None:
                    mother_name = tr_capitalize(stripper(str(active_page.cell(row, 9).value)))

                birthdate = ""
                if active_page.cell(row, 10).value is not None:
                    birthdate = str(active_page.cell(row, 10).value)[:10]

                birth_place = ""
                if active_page.cell(row, 11).value is not None:
                    birth_place = tr_capitalize(stripper(str(active_page.cell(row, 11).value)))

                self_phone = ""
                if active_page.cell(row, 12).value is not None:
                    self_phone = str(active_page.cell(row, 1).value)

                father_phone = ""
                if active_page.cell(row, 13).value is not None:
                    father_phone = str(active_page.cell(row, 13).value)

                mother_phone = ""
                if active_page.cell(row, 14).value is not None:
                    mother_phone = str(active_page.cell(row, 14).value)

                email = ""
                if active_page.cell(row, 15).value is not None:
                    email = str(active_page.cell(row, 15).value)

                photo_address = ""
                if active_page.cell(row, 16).value is not None:
                    photo_address = str(active_page.cell(row, 16).value)

                gender = active_page.cell(row, 17).value

                wp_name = active_page.cell(row, 18).value

                starting_date = ""
                if active_page.cell(row, 19).value is not None:
                    starting_date = str(active_page.cell(row, 19).value)[:10]

                gc = active_page.cell(row, 20).value

                control_list = [identity_number_control(id_number, active_student_id_list), name_control(name), surname_control(surname), number_control(number),
                                archive_identity_number_control(id_number, passive_student_id_list), branch_name is not None,
                                gender is not None, wp_name is not None, gc is not None, is_date(recorded_date), is_date(birthdate), is_date(starting_date)]

                if False in control_list:
                    QMessageBox.warning(None, "Uyarı", str(row) + ". satırda hata var!\ndüzeltip yeniden yükleyiniz.")
                    new_student_list.clear()
                    break

                class_id = "NULL"
                branch_department_id = self.connection.find(f"""SELECT departmentID FROM branches WHERE name="{branch_name}" """)
                if class_name != "":
                    class_id = self.connection.find(f"""SELECT id FROM classes WHERE name="{class_name}" """)
                    class_department_id = self.connection.find(f"""SELECT departmentID FROM classes WHERE name="{class_name}" """)
                    if class_department_id != branch_department_id:
                        QMessageBox.warning(None, "Uyarı", str(row) + ". satırda seçtiğiniz sınıf seçtiğiniz branşa ait değil.")
                        new_student_list.clear()
                        break

                if wp_name != "Okul":
                    wp_department_id = self.connection.find(f"""SELECT departmentID FROM workplaces WHERE name="{wp_name}" """)
                    if wp_department_id != branch_department_id:
                        QMessageBox.warning(None, "Uyarı", str(row) + ". satırda seçtiğiniz branş seçtiğiniz işletmeye ait değil.")
                        new_student_list.clear()
                        break

                branch_id = self.connection.find(f"""SELECT id FROM branches WHERE name="{branch_name}" """)

                new_student_list.append({"id_number": id_number, "name": name, "surname": surname, "number": number, "branch_id": branch_id, "class_id": class_id, "record_date": recorded_date,
                                         "father": father_name, "mother": mother_name, "birth_date": birthdate, "birth_place": birth_place, "workplace_name": wp_name, "starting_date": starting_date,
                                         "self_phone": self_phone, "father_phone": father_phone, "mother_phone": mother_phone, "email": email, "photo_address": photo_address, "gender": gender})

                active_student_id_list.append(id_number)

        for item in new_student_list:
            print(item)
            self.connection.recorder(f"""INSERT INTO students VALUES("{item["id_number"]}", "{item["name"]}", "{item["surname"]}", "{item["number"]}", {item["branch_id"]}, {item["class_id"]},
            "{item["record_date"]}", "{item["father"]}", "{item["mother"]}", "{item["birth_date"]}", "{item["birth_place"]}", "E", "{item["self_phone"]}", "{item["father_phone"]}",
            "{item["mother_phone"]}", "{item["email"]}", "{item["photo_address"]}", "E","{item["gender"]}") """)
            if item["workplace_name"] == "Okul":
                self.connection.updater(f"""UPDATE students SET is_going="H" WHERE id_number="{item["id_number"]}" """)
                self.connection.recorder(f""" INSERT INTO temp_workplace(studentID, staffID, starting_date) VALUES ("{item["id_number"]}", NULL, "{item["starting_date"]}") """)
            else:
                wp_id = self.connection.find(f""" SELECT id FROM workplaces WHERE name = "{item["workplace_name"]}" """)
                self.connection.recorder(f""" INSERT INTO history (studentID, workplaceID, starting_date) VALUES ("{item["id_number"]}", {wp_id}, "{item["starting_date"]}") """)
                number_of_student_in_new_wp = self.connection.selector(f""" SELECT COUNT(studentID) FROM history  WHERE workplaceID = {wp_id} AND leaving_date IS NULL """)[0][0]
                if number_of_student_in_new_wp > 1:
                    self.settle_workplace(item["workplace_name"], item["id_number"])

    def settle_workplace(self, p_wp_name, p_id_number):
        wp_day_id, wp_staff_id = self.connection.selector(f""" SELECT dayID, staffID FROM workplaces WHERE name = "{p_wp_name}" """)[0]
        class_id = self.connection.find(f"""SELECT classID FROM students WHERE id_number = "{p_id_number}" """)
        if class_id is not None:
            class_days_id_list = [day[0] for day in self.connection.selector(f""" SELECT dayID FROM classes_days WHERE classID={class_id}""")]
        else:
            class_days_id_list = [day[0] for day in self.connection.selector(f""" SELECT dayID FROM classes_days""")]
        if wp_day_id is not None and wp_staff_id is not None:
            if wp_day_id not in class_days_id_list:
                day_name = self.connection.find(f""" SELECT name FROM days WHERE id = {wp_day_id} """)
                staff_name = " ".join(self.connection.selector(f""" SELECT name, surname FROM staffs WHERE id_number = "{wp_staff_id}" """)[0])
                self.connection.updater(f""" UPDATE workplaces  SET staffID = NULL, dayID=NULL WHERE name = "{p_wp_name}" """)
                message_box(f"""{staff_name}'in {day_name} günü {p_wp_name} işletmesindeki görevi sıfırlanmıştır.""")

    def load_workplaces_file(self):
        path, ok = QFileDialog.getOpenFileName(self, caption="Aç", directory=os.getcwd(), filter="Excel Files(*.xlsx)")
        if ok:
            wb = load_workbook(path, read_only=True)
            if "isletme" in wb.sheetnames:
                self.read_wp_excel_file(path)
            else:
                QMessageBox.warning(None, "Uyarı", "dosya içinde gerekli sayfalar yok sayfayı yeniden indiriniz.")

    def load_students_file(self):
        path, ok = QFileDialog.getOpenFileName(self, caption="Aç", directory=os.getcwd(), filter="Excel Files(*.xlsx)")
        if ok:
            wb = load_workbook(path, read_only=True)
            if "Ogrenciler" in wb.sheetnames:
                self.read_student_excel_file(path)
            else:
                QMessageBox.warning(None, "Uyarı", "dosya içinde gerekli sayfalar yok sayfayı yeniden indiriniz.")

    def save_excel_for_wp(self):
        if not os.path.isdir(os.path.join(os.getcwd(), "helper_files")):
            os.mkdir(os.path.join(os.getcwd(), "helper_files"))

        title_list = ["İşyeri Adı*", "İşyeri Sahibi*", "Alan Adı*", "Usta Öğretici*", "Mahalle Adı*", "adres satırı 1", "adres satırı 2", "telefon1", "telefon2", "email", "devlet desteği*"]
        neighborhoods_list = self.connection.selector(f"""SELECT * FROM neighborhoods""")
        departments_list = self.connection.selector(f"""SELECT * FROM departments ORDER BY name""")
        wp_list = self.connection.selector(f"""SELECT * FROM workplaces""")

        excel_path = os.path.join(os.getcwd(), "helper_files/workplaces.xlsx")

        wb = Workbook()

        wb.create_sheet("isletme", 0)
        active_page = wb["isletme"]
        active_page.append(title_list)

        active_page.cell(row=1, column=25, value="id")
        active_page.cell(row=1, column=26, value="Alan Adı")
        for i in range(2, len(departments_list)+2):
            active_page.cell(row=i, column=25, value=departments_list[i-2][0])
            active_page.cell(row=i, column=26, value=departments_list[i-2][1])

        active_page.cell(row=1, column=28, value="id")
        active_page.cell(row=1, column=29, value="Mahalle Adı")
        for i in range(2, len(neighborhoods_list) + 2):
            active_page.cell(row=i, column=28, value=neighborhoods_list[i - 2][0])
            active_page.cell(row=i, column=29, value=neighborhoods_list[i - 2][1])

        active_page.cell(row=1, column=31, value="id")
        active_page.cell(row=1, column=32, value="Varolan İşletme Adı")
        for i in range(2, len(wp_list) + 2):
            active_page.cell(row=i, column=31, value=wp_list[i - 2][0])
            active_page.cell(row=i, column=32, value=wp_list[i - 2][1])

        del wb["Sheet"]

        dv_depart = DataValidation(type="list", formula1=f'=$Z$2:$Z${len(departments_list) + 1}', allow_blank=False)
        dv_neigh = DataValidation(type="list", formula1=f'=$AC$2:$AC${len(neighborhoods_list) + 1}', allow_blank=False)
        dv_yes_no = DataValidation(type="list", formula1='"E,H"', allow_blank=False)
        active_page.add_data_validation(dv_neigh)
        active_page.add_data_validation(dv_depart)
        active_page.add_data_validation(dv_yes_no)
        dv_depart.add("C2:C10000")
        dv_neigh.add("E2:E10000")
        dv_yes_no.add("K2:K10000")

        wb.save(excel_path)
        wb.close()

        wb = load_workbook(excel_path)
        path, ok = QFileDialog.getSaveFileName(self, caption="Kaydet", directory=excel_path, filter="Excel files (*.xlsx)")
        if ok:
            wb.save(path)

    def save_excel_file_for_student(self):
        if not os.path.isdir(os.path.join(os.getcwd(), "helper_files")):
            os.mkdir(os.path.join(os.getcwd(), "helper_files"))
        title_list = ["TC kimlik No", "Ad", "Soyad", "Okul No", "Dal", "Sınıf", "Okula Kayıt Tarihi", "Baba Adı", "Ana Adı", "Doğum Tarihi", "Doğum Yeri",  "Kendi Telefonu",
                      "Baba Telefonu", "Ana Telefonu", "email", "foto adresi", "Cinsiyet", "İşletme Adı", "İşe Giriş Tarihi", "Devlet Katkısı İstiyor"]
        branch_list = [branch[0] for branch in self.connection.selector(f"""SELECT name FROM branches""")]
        class_list = [""] + [classroom[0] for classroom in self.connection.selector(f"""SELECT name FROM classes""")]
        wp_list = ["Okul"] + [wp[0] for wp in self.connection.selector(f"""SELECT name FROM workplaces""")]
        print(wp_list)

        excel_path = os.path.join(os.getcwd(), "helper_files/students.xlsx")

        wb = Workbook()

        wb.create_sheet("Ogrenciler", 0)
        del wb["Sheet"]
        active_page = wb["Ogrenciler"]
        active_page.append(title_list)

        active_page.cell(row=1, column=30, value="Dal Adı")
        for i in range(2, len(branch_list) + 2):
            active_page.cell(row=i, column=30, value=branch_list[i - 2])

        active_page.cell(row=1, column=31, value="Sınıf Adı")
        for i in range(2, len(class_list) + 2):
            active_page.cell(row=i, column=31, value=class_list[i - 2])

        active_page.cell(row=1, column=32, value="Varolan İşletme Adı")
        for i in range(2, len(wp_list) + 2):
            active_page.cell(row=i, column=32, value=wp_list[i - 2])

        dv_branch = DataValidation(type="list", formula1=f'=$AD$2:$AD${len(branch_list) + 1}', allow_blank=False)
        dv_class = DataValidation(type="list", formula1=f'=$AE$2:$AE${len(class_list) + 1}', allow_blank=False)
        dv_wp = DataValidation(type="list", formula1=f'=$AF$2:$AF${len(wp_list) + 1}', allow_blank=False)
        dv_gender = DataValidation(type="list", formula1='"E,K"', allow_blank=False)
        dv_gc = DataValidation(type="list", formula1='"E,H"', allow_blank=False)
        active_page.add_data_validation(dv_branch)
        active_page.add_data_validation(dv_class)
        active_page.add_data_validation(dv_wp)
        active_page.add_data_validation(dv_gender)
        active_page.add_data_validation(dv_gc)
        dv_branch.add("E2:E10000")
        dv_class.add("F2:F10000")
        dv_gender.add("Q2:Q10000")
        dv_wp.add("R2:R10000")
        dv_gc.add("T2:T10000")

        wb.save(excel_path)
        wb.close()

        wb = load_workbook(excel_path)
        path, ok = QFileDialog.getSaveFileName(self, caption="Kaydet", directory=excel_path, filter="Excel files (*.xlsx)")
        if ok:
            wb.save(path)

    def closeEvent(self, event):
        self.connection.db_closer()


if __name__ == "__main__":
    app = QApplication([])
    window = App()
    window.show()
    sys.exit(app.exec())
